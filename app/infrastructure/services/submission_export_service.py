import csv
from io import BytesIO, StringIO
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.domain.models import FormSubmission
from app.domain.services.submission_export_service import ISubmissionExportService, ExportFormat


class SubmissionExportService(ISubmissionExportService):
    """Service for exporting submissions to various formats."""
    
    # Translations
    TRANSLATIONS = {
        "en": {
            "form_title": "Form Title",
            "submitted_by": "Submitted By",
            "email": "Email",
            "submitted_at": "Submitted At",
            "field": "Field",
            "value": "Value",
            "files_uploaded": "file(s) uploaded",
            "no_files": "No files",
            "signature_present": "[Digital Signature Present]",
            "no_signature": "No signature",
            "n_a": "N/A",
        },
        "uk": {
            "form_title": "Назва форми",
            "submitted_by": "Подано",
            "email": "Електронна пошта",
            "submitted_at": "Дата подання",
            "field": "Поле",
            "value": "Значення",
            "files_uploaded": "файл(ів) завантажено",
            "no_files": "Немає файлів",
            "signature_present": "[Цифровий підпис присутній]",
            "no_signature": "Немає підпису",
            "n_a": "Н/Д",
        }
    }
    
    async def export_submission(
        self, 
        submission: FormSubmission, 
        format: ExportFormat,
        locale: str = "en"
    ) -> tuple[BytesIO, str]:
        """Export submission to specified format."""
        if format == "csv":
            return await self._export_to_csv(submission, locale)
        elif format == "xlsx":
            return await self._export_to_xlsx(submission, locale)
        else:
            raise ValueError(f"Unsupported export format: {format}")
    
    def _get_text(self, key: str, locale: str) -> str:
        """Get translated text."""
        return self.TRANSLATIONS.get(locale, self.TRANSLATIONS["en"]).get(key, key)
    
    async def _export_to_csv(self, submission: FormSubmission, locale: str) -> tuple[BytesIO, str]:
        """Export submission to CSV format."""
        # Create CSV in memory
        output = StringIO()
        writer = csv.writer(output)
        
        # Add metadata
        writer.writerow([self._get_text("form_title", locale), submission.form.title if submission.form else "Unknown"])
        writer.writerow([self._get_text("submitted_by", locale), submission.user.name if submission.user else "Unknown"])
        writer.writerow([self._get_text("email", locale), submission.user.email if submission.user and submission.user.email else self._get_text("n_a", locale)])
        writer.writerow([self._get_text("submitted_at", locale), submission.submitted_at.strftime("%Y-%m-%d %H:%M:%S")])
        writer.writerow([])  # Empty row
        
        # Add headers
        writer.writerow([self._get_text("field", locale), self._get_text("value", locale)])
        
        # Add field values
        if submission.form and submission.form.fields:
            sorted_fields = sorted(submission.form.fields, key=lambda f: f.order)
            for field in sorted_fields:
                field_value = next(
                    (fv for fv in submission.field_values if fv.field_id == field.id), 
                    None
                )
                
                value = ""
                if field.field_type == "files":
                    # For file fields, show count instead of names
                    files_count = len([
                        f for f in submission.files 
                        if f.field_id == field.id
                    ])
                    if files_count > 0:
                        value = f"{files_count} file(s) uploaded"
                    else:
                        value = "No files"
                elif field.field_type == "signature":
                    # For signature fields, don't show base64 data
                    if field_value and field_value.value:
                        value = "[Digital Signature Present]"
                    else:
                        value = "No signature"
                elif field_value:
                    value = field_value.value or ""
                
                writer.writerow([field.label, value])
        
        # Convert to bytes
        csv_content = output.getvalue()
        output.close()
        
        buffer = BytesIO(csv_content.encode('utf-8-sig'))  # UTF-8 with BOM for Excel compatibility
        buffer.seek(0)
        
        # Generate filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        form_title = submission.form.title if submission.form else "submission"
        user_name = submission.user.name if submission.user and submission.user.name else "user"
        # Sanitize filename parts
        def sanitize(text: str) -> str:
            return "".join(c if c.isalnum() or c in (' ', '-', '_') else '_' for c in text)
        safe_title = sanitize(form_title)
        safe_user = sanitize(user_name)
        filename = f"{safe_title}_{safe_user}_{timestamp}.csv"
        
        return buffer, filename
    
    async def _export_to_xlsx(self, submission: FormSubmission, locale: str) -> tuple[BytesIO, str]:
        """Export submission to XLSX format."""
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Submission"
        
        # Define styles
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        metadata_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
        metadata_font = Font(bold=True, size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        row = 1
        
        # Add metadata section
        metadata = [
            (self._get_text("form_title", locale), submission.form.title if submission.form else "Unknown"),
            (self._get_text("submitted_by", locale), submission.user.name if submission.user else "Unknown"),
            (self._get_text("email", locale), submission.user.email if submission.user and submission.user.email else self._get_text("n_a", locale)),
            (self._get_text("submitted_at", locale), submission.submitted_at.strftime("%Y-%m-%d %H:%M:%S")),
        ]
        
        for label, value in metadata:
            cell_label = ws.cell(row=row, column=1, value=label)
            cell_value = ws.cell(row=row, column=2, value=value)
            cell_label.fill = metadata_fill
            cell_label.font = metadata_font
            cell_label.border = border
            cell_value.border = border
            row += 1
        
        # Empty row
        row += 1
        
        # Add headers
        header_cell_1 = ws.cell(row=row, column=1, value=self._get_text("field", locale))
        header_cell_2 = ws.cell(row=row, column=2, value=self._get_text("value", locale))
        
        for cell in [header_cell_1, header_cell_2]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        row += 1
        
        # Add field values
        if submission.form and submission.form.fields:
            sorted_fields = sorted(submission.form.fields, key=lambda f: f.order)
            for field in sorted_fields:
                field_value = next(
                    (fv for fv in submission.field_values if fv.field_id == field.id), 
                    None
                )
                
                value = ""
                if field.field_type == "files":
                    # For file fields, show count instead of names
                    files_count = len([
                        f for f in submission.files 
                        if f.field_id == field.id
                    ])
                    if files_count > 0:
                        value = f"{files_count} {self._get_text('files_uploaded', locale)}"
                    else:
                        value = self._get_text("no_files", locale)
                elif field.field_type == "signature":
                    # For signature fields, don't show base64 data
                    if field_value and field_value.value:
                        value = self._get_text("signature_present", locale)
                    else:
                        value = self._get_text("no_signature", locale)
                elif field_value:
                    value = field_value.value or ""
                
                cell_field = ws.cell(row=row, column=1, value=field.label)
                cell_value = ws.cell(row=row, column=2, value=value)
                
                cell_field.border = border
                cell_value.border = border
                cell_value.alignment = Alignment(wrap_text=True, vertical='top')
                
                row += 1
        
        # Adjust column widths
        ws.column_dimensions[get_column_letter(1)].width = 30
        ws.column_dimensions[get_column_letter(2)].width = 50
        
        # Save to buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        # Generate filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        form_title = submission.form.title if submission.form else "submission"
        user_name = submission.user.name if submission.user and submission.user.name else "user"
        # Sanitize filename parts
        def sanitize(text: str) -> str:
            return "".join(c if c.isalnum() or c in (' ', '-', '_') else '_' for c in text)
        safe_title = sanitize(form_title)
        safe_user = sanitize(user_name)
        filename = f"{safe_title}_{safe_user}_{timestamp}.xlsx"
        
        return buffer, filename
